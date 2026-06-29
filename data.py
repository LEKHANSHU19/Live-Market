"""
data.py — Quantitative Risk & Derivatives Analytics Platform
============================================================
Market data ingestion layer.

Responsibilities
----------------
    load_price_data      Read a price series for one ticker from an Excel file,
                         returning a clean, sorted, datetime-indexed pd.Series.
    compute_log_returns  Compute daily log returns ln(P_t / P_{t-1}) from a
                         price series, dropping the leading NaN.

Design principles
-----------------
- Both functions are pure: they accept inputs, return outputs, and have no
  side effects (no global state, no in-place mutation).
- All validation raises descriptive exceptions so callers always know exactly
  what went wrong.
- No hardcoded paths or ticker names: those come from config.MARKET.

Source mapping
--------------
    load_price_data     ← E1 Cell 4  (file search + pd.read_excel + ^GSPC extract)
    compute_log_returns ← E1 Cell 4  (np.log(prices / prices.shift(1)).dropna())

Imports used by downstream modules
-----------------------------------
    risk.py      calls both functions to build the backtesting dataset
    main.py      calls load_price_data with MARKET.DATA_FILEPATH
"""

from __future__ import annotations

import os
from pathlib import Path

import numpy as np
import pandas as pd

from quant_platform.config import MARKET


# ── Public API ────────────────────────────────────────────────────────────────

def load_price_data(
    filepath: str | os.PathLike[str],
    ticker: str = MARKET.SP500_TICKER,
) -> pd.Series:
    """Load an adjusted-close price series from an Excel workbook.

    The workbook is expected to contain at least two columns: one named
    ``"Date"`` and one named after *ticker*.  All other columns are ignored.
    The function cleans, sorts, and validates the series before returning it.

    Parameters
    ----------
    filepath:
        Absolute or relative path to the ``.xlsx`` file.  The file must be
        readable by ``openpyxl``.
    ticker:
        Column name in the workbook that holds the price series.
        Defaults to ``MARKET.SP500_TICKER`` (``"^GSPC"``).

    Returns
    -------
    pd.Series
        Daily close prices with a ``DatetimeIndex``, sorted ascending, NaNs
        dropped, and the series named after *ticker*.

    Raises
    ------
    FileNotFoundError
        If *filepath* does not exist on disk.
    ValueError
        If the workbook has no ``"Date"`` column, no column matching *ticker*,
        or if the resulting series is empty after dropping NaNs.
    TypeError
        If *filepath* is not a string or path-like object.

    Examples
    --------
    >>> from config import MARKET
    >>> prices = load_price_data(MARKET.DATA_FILEPATH, MARKET.SP500_TICKER)
    >>> print(prices.head())
    Date
    2020-01-02    3257.85
    2020-01-03    3234.85
    ...
    Name: ^GSPC, dtype: float64
    """
    # ── Input validation ─────────────────────────────────────────────────────
    if not isinstance(filepath, (str, Path, os.PathLike)):
        raise TypeError(
            f"filepath must be a string or path-like object, "
            f"got {type(filepath).__name__!r}."
        )

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(
            f"Data file not found: '{path.resolve()}'\n"
            f"Place '{path.name}' in '{path.parent.resolve()}' and retry."
        )
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError(
            f"Expected an Excel file (.xlsx / .xls), got '{path.suffix}'."
        )
    if not isinstance(ticker, str) or not ticker.strip():
        raise ValueError(
            f"ticker must be a non-empty string, got {ticker!r}."
        )

    # ── Read workbook ────────────────────────────────────────────────────────
    try:
        raw: pd.DataFrame = pd.read_excel(path, engine="openpyxl")
    except Exception as exc:
        raise ValueError(
            f"Could not read '{path.name}' with openpyxl: {exc}"
        ) from exc

    # ── Column validation ────────────────────────────────────────────────────
    if "Date" not in raw.columns:
        raise ValueError(
            f"Expected a column named 'Date' in '{path.name}'. "
            f"Found columns: {list(raw.columns)}."
        )
    if ticker not in raw.columns:
        raise ValueError(
            f"Ticker '{ticker}' not found in '{path.name}'. "
            f"Available columns: {[c for c in raw.columns if c != 'Date']}."
        )

    # ── Clean and index ──────────────────────────────────────────────────────
    raw["Date"] = pd.to_datetime(raw["Date"], errors="coerce")

    n_bad_dates = raw["Date"].isna().sum()
    if n_bad_dates > 0:
        # Warn but don't fail: drop unparseable rows rather than crashing
        import warnings
        warnings.warn(
            f"{n_bad_dates} row(s) in '{path.name}' had unparseable dates "
            f"and were dropped.",
            stacklevel=2,
        )

    prices: pd.Series = (
        raw.dropna(subset=["Date"])
        .set_index("Date")
        .sort_index()
        [ticker]
        .dropna()
        .rename(ticker)
        .astype(float)
    )

    if prices.empty:
        raise ValueError(
            f"Price series for '{ticker}' in '{path.name}' is empty after "
            f"dropping NaN values. Check the data file."
        )

    # ── Sanity checks ────────────────────────────────────────────────────────
    n_non_positive = (prices <= 0).sum()
    if n_non_positive > 0:
        import warnings
        warnings.warn(
            f"{n_non_positive} non-positive price(s) detected in '{ticker}'. "
            f"Log-return computation will produce NaN or -inf for those dates.",
            stacklevel=2,
        )

    return prices


def compute_log_returns(prices: pd.Series) -> pd.Series:
    """Compute daily log returns from a price series.

    Applies the transformation:

        r_t = ln(P_t / P_{t-1})

    which is equivalent to ``ln(P_t) - ln(P_{t-1})``.  The leading NaN
    produced by the shift is dropped, so the returned series has
    ``len(prices) - 1`` observations.

    Parameters
    ----------
    prices:
        A ``pd.Series`` of positive prices with a monotonically increasing
        ``DatetimeIndex``.  Typically the output of :func:`load_price_data`.

    Returns
    -------
    pd.Series
        Daily log returns with the same index type as *prices* (minus the
        first date), named ``"log_return"``.

    Raises
    ------
    TypeError
        If *prices* is not a ``pd.Series``.
    ValueError
        If *prices* has fewer than 2 observations, contains non-positive
        values, or produces an all-NaN return series.

    Notes
    -----
    The log-return formulation is preferred over the simple return
    ``P_t / P_{t-1} - 1`` for two reasons that matter in the risk module:

    1. Log returns are time-additive: a 10-day log return equals the sum
       of 10 daily log returns — this is exploited by the ``sqrt(10)``
       VaR scaling rule in ``risk.py``.
    2. They are symmetric around zero, which is consistent with the
       normal distribution assumption used for parametric VaR.

    Examples
    --------
    >>> from config import MARKET
    >>> prices = load_price_data(MARKET.DATA_FILEPATH, MARKET.SP500_TICKER)
    >>> returns = compute_log_returns(prices)
    >>> print(f"Observations : {len(returns)}")
    >>> print(f"Mean daily r : {returns.mean():.6f}")
    >>> print(f"Daily sigma  : {returns.std():.6f}")
    """
    # ── Input validation ─────────────────────────────────────────────────────
    if not isinstance(prices, pd.Series):
        raise TypeError(
            f"prices must be a pd.Series, got {type(prices).__name__!r}."
        )
    if len(prices) < 2:
        raise ValueError(
            f"prices must have at least 2 observations to compute returns; "
            f"got {len(prices)}."
        )

    n_non_positive = (prices <= 0).sum()
    if n_non_positive > 0:
        raise ValueError(
            f"prices contains {n_non_positive} non-positive value(s). "
            f"Log returns are undefined for prices ≤ 0."
        )

    # ── Computation ──────────────────────────────────────────────────────────
    # Identical to E1 Cell 4: np.log(prices / prices.shift(1)).dropna()
    log_returns: pd.Series = np.log(prices / prices.shift(1)).dropna()
    log_returns.name = "log_return"

    # ── Post-computation validation ──────────────────────────────────────────
    if log_returns.isna().all():
        raise ValueError(
            "All computed log returns are NaN. "
            "Check that prices is a valid, non-constant price series."
        )

    return log_returns


# ── Convenience summary ───────────────────────────────────────────────────────

def summarise_price_data(
    prices: pd.Series,
    returns: pd.Series,
) -> None:
    """Print a concise loading summary matching the E1 Cell 4 console output.

    Intended for use in ``main.py`` immediately after loading data, so the
    user can verify the series at a glance.

    Parameters
    ----------
    prices:
        Output of :func:`load_price_data`.
    returns:
        Output of :func:`compute_log_returns`.
    """
    from scipy.stats import norm  # local import — only needed for display

    z_alpha = norm.ppf(1.0 - 0.99)  # -2.3263 at 99 % confidence

    ticker = prices.name or "unknown"
    print("=" * 55)
    print("DATA LOADED SUCCESSFULLY")
    print("=" * 55)
    print(f"  Ticker        : {ticker}")
    print(f"  From          : {prices.index[0].date()}")
    print(f"  To            : {prices.index[-1].date()}")
    print(f"  Trading days  : {len(prices):,}")
    print(f"  First price   : {prices.iloc[0]:,.2f}")
    print(f"  Last price    : {prices.iloc[-1]:,.2f}")
    print()
    print(f"  Daily returns : {len(returns):,} observations")
    print(f"  Full-sample σ : {returns.std() * 100:.4f} % per day")
    print(f"  z_α (99% VaR) : {z_alpha:.4f}")
    print()


# ── Example usage (run directly: python data.py) ──────────────────────────────

if __name__ == "__main__":
    # Usage example — replace DATA_FILEPATH in config.py with your actual path,
    # or pass an explicit path here for a quick smoke-test.
    #
    #   python data.py
    #
    example_path = MARKET.DATA_FILEPATH

    print(f"Loading data from : {example_path!r}")
    print(f"Ticker            : {MARKET.SP500_TICKER!r}")
    print()

    try:
        prices = load_price_data(example_path, MARKET.SP500_TICKER)
        returns = compute_log_returns(prices)
        summarise_price_data(prices, returns)

        print("Return series (first 5 rows):")
        print(returns.head().to_string())
        print()
        print("Return series (last 5 rows):")
        print(returns.tail().to_string())

    except FileNotFoundError as exc:
        print(f"[FileNotFoundError] {exc}")
    except ValueError as exc:
        print(f"[ValueError] {exc}")


# ── Unit test section ─────────────────────────────────────────────────────────
# Run with:  python -m pytest data.py -v
# These tests use only synthetic data and have no file-system dependencies.

def _make_prices(values: list[float], start: str = "2024-01-02") -> pd.Series:
    """Helper: build a DatetimeIndex price Series from a plain list."""
    idx = pd.date_range(start=start, periods=len(values), freq="B")
    return pd.Series(values, index=idx, name="^TEST", dtype=float)


class TestLoadPriceData:
    """Validation and edge-case tests for load_price_data."""

    def test_raises_on_missing_file(self) -> None:
        import pytest
        with pytest.raises(FileNotFoundError, match="not found"):
            load_price_data("nonexistent_file.xlsx")

    def test_raises_on_wrong_extension(self, tmp_path) -> None:
        import pytest
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("Date,^GSPC\n2024-01-02,100\n")
        with pytest.raises(ValueError, match="Expected an Excel file"):
            load_price_data(csv_file)

    def test_raises_on_non_path_input(self) -> None:
        import pytest
        with pytest.raises(TypeError, match="path-like"):
            load_price_data(12345)  # type: ignore[arg-type]

    def test_raises_on_missing_date_column(self, tmp_path) -> None:
        import pytest
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Price"])
        ws.append([100.0])
        path = tmp_path / "no_date.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="'Date'"):
            load_price_data(path, "Price")

    def test_raises_on_missing_ticker_column(self, tmp_path) -> None:
        import pytest
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "WRONG"])
        ws.append(["2024-01-02", 100.0])
        path = tmp_path / "wrong_ticker.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="not found"):
            load_price_data(path, "^GSPC")

    def test_returns_sorted_datetime_series(self, tmp_path) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "^GSPC"])
        # deliberately reversed order to test sort
        ws.append(["2024-01-05", 105.0])
        ws.append(["2024-01-02", 100.0])
        ws.append(["2024-01-03", 102.0])
        path = tmp_path / "prices.xlsx"
        wb.save(path)
        result = load_price_data(path, "^GSPC")
        assert isinstance(result.index, pd.DatetimeIndex)
        assert result.index.is_monotonic_increasing
        assert result.iloc[0] == 100.0   # first after sort
        assert result.iloc[-1] == 105.0  # last after sort
        assert result.name == "^GSPC"

    def test_drops_nan_rows(self, tmp_path) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "^GSPC"])
        ws.append(["2024-01-02", 100.0])
        ws.append(["2024-01-03", None])   # NaN price row
        ws.append(["2024-01-04", 102.0])
        path = tmp_path / "with_nan.xlsx"
        wb.save(path)
        result = load_price_data(path, "^GSPC")
        assert len(result) == 2           # NaN row was dropped
        assert result.isna().sum() == 0


class TestComputeLogReturns:
    """Mathematical correctness and edge-case tests for compute_log_returns."""

    def test_length_is_n_minus_one(self) -> None:
        prices = _make_prices([100.0, 101.0, 99.0, 102.0, 100.5])
        returns = compute_log_returns(prices)
        assert len(returns) == len(prices) - 1

    def test_no_leading_nan(self) -> None:
        prices = _make_prices([100.0, 101.0, 102.0])
        returns = compute_log_returns(prices)
        assert returns.isna().sum() == 0

    def test_known_value(self) -> None:
        import math
        prices = _make_prices([100.0, math.e * 100.0])  # P1/P0 = e  => ln = 1.0
        returns = compute_log_returns(prices)
        assert abs(returns.iloc[0] - 1.0) < 1e-12

    def test_series_name_is_log_return(self) -> None:
        prices = _make_prices([100.0, 101.0])
        returns = compute_log_returns(prices)
        assert returns.name == "log_return"

    def test_raises_on_non_series(self) -> None:
        import pytest
        with pytest.raises(TypeError, match="pd.Series"):
            compute_log_returns([100.0, 101.0])  # type: ignore[arg-type]

    def test_raises_on_single_observation(self) -> None:
        import pytest
        prices = _make_prices([100.0])
        with pytest.raises(ValueError, match="at least 2"):
            compute_log_returns(prices)

    def test_raises_on_non_positive_prices(self) -> None:
        import pytest
        prices = _make_prices([100.0, 0.0, 101.0])
        with pytest.raises(ValueError, match="non-positive"):
            compute_log_returns(prices)

    def test_zero_return_for_flat_series(self) -> None:
        prices = _make_prices([100.0, 100.0, 100.0, 100.0])
        returns = compute_log_returns(prices)
        assert (returns == 0.0).all()

    def test_index_alignment_with_prices(self) -> None:
        prices = _make_prices([100.0, 101.0, 102.0, 103.0])
        returns = compute_log_returns(prices)
        # Returns index should be prices.index[1:] exactly
        assert (returns.index == prices.index[1:]).all()

    def test_symmetry_property(self) -> None:
        # ln(P1/P0) + ln(P0/P1) should equal zero (log-return symmetry)
        import math
        up   = compute_log_returns(_make_prices([100.0, 110.0]))
        down = compute_log_returns(_make_prices([110.0, 100.0]))
        assert abs(up.iloc[0] + down.iloc[0]) < 1e-12

    def test_additivity_over_horizon(self) -> None:
        # Two-period log return == sum of two one-period log returns
        import math
        prices = _make_prices([100.0, 105.0, 110.25])
        returns = compute_log_returns(prices)
        two_period = math.log(110.25 / 100.0)
        assert abs(returns.sum() - two_period) < 1e-12
